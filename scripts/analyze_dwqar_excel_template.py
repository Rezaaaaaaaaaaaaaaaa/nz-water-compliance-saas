"""
DWQAR Excel Template Analyzer
Analyzes the official DWQAR_Reporting_Template.xlsx to extract field requirements
"""

import openpyxl
import json
from pathlib import Path

TEMPLATE_PATH = Path("C:/compliance-saas/docs/regulations/taumata-arowai/DWQAR_Reporting_Template.xlsx")
OUTPUT_PATH = Path("C:/compliance-saas/docs/regulatory-analysis/dwqar_excel_field_mapping.json")

def analyze_excel_template():
    """Analyze the DWQAR Excel template structure"""

    print("=" * 70)
    print("DWQAR EXCEL TEMPLATE ANALYSIS")
    print("=" * 70)
    print()

    if not TEMPLATE_PATH.exists():
        print(f"[ERROR] Template not found: {TEMPLATE_PATH}")
        return

    # Load workbook
    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    print(f"Workbook: {TEMPLATE_PATH.name}")
    print(f"Total sheets: {len(wb.sheetnames)}\n")

    analysis = {
        "template_file": str(TEMPLATE_PATH.name),
        "analysis_date": "2025-10-05",
        "sheets": []
    }

    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        print(f"\n{'=' * 70}")
        print(f"SHEET: {sheet_name}")
        print(f"{'=' * 70}")

        sheet = wb[sheet_name]

        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column

        print(f"Dimensions: {max_row} rows x {max_col} columns\n")

        # Extract headers (usually in first few rows)
        headers = []
        data_start_row = None

        # Look for headers in first 20 rows
        for row_idx in range(1, min(21, max_row + 1)):
            row_values = []
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                row_values.append(str(cell.value) if cell.value is not None else "")

            # Check if this row has significant content
            non_empty = [v for v in row_values if v.strip()]
            if len(non_empty) > 3:  # Row with multiple headers
                print(f"Row {row_idx}: {non_empty[:10]}...")  # Show first 10
                if not data_start_row and any(keyword in ' '.join(non_empty).lower() for keyword in ['supply', 'name', 'code', 'date', 'water', 'compliance']):
                    data_start_row = row_idx + 1
                    headers = row_values
                    print(f"  ^--- Identified as HEADER ROW (data starts at row {data_start_row})")

        print()

        # Extract full header details
        if headers:
            clean_headers = [h.strip() for h in headers if h.strip()]
            print(f"Found {len(clean_headers)} header columns:\n")

            header_details = []
            for idx, header in enumerate(clean_headers[:30], 1):  # Show first 30
                print(f"  {idx}. {header}")
                header_details.append({
                    "column_index": idx,
                    "header_name": header,
                    "database_mapping": f"TO_BE_MAPPED_{idx}",
                    "required": "UNKNOWN",
                    "data_type": "UNKNOWN"
                })

        print()

        # Sample some data rows
        if data_start_row and data_start_row <= max_row:
            print(f"\nSample data (rows {data_start_row} to {min(data_start_row + 3, max_row)}):")
            for row_idx in range(data_start_row, min(data_start_row + 4, max_row + 1)):
                row_data = []
                for col_idx in range(1, min(11, max_col + 1)):  # First 10 columns
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    row_data.append(str(cell.value) if cell.value is not None else "")
                print(f"  Row {row_idx}: {row_data}")

        # Store sheet analysis
        sheet_analysis = {
            "sheet_name": sheet_name,
            "dimensions": {
                "rows": max_row,
                "columns": max_col
            },
            "header_row": 1 if headers else None,
            "data_start_row": data_start_row,
            "headers": header_details if headers else [],
            "notes": []
        }

        # Add notes based on sheet name
        if "supply" in sheet_name.lower():
            sheet_analysis["notes"].append("Contains water supply information")
        if "test" in sheet_name.lower() or "quality" in sheet_name.lower():
            sheet_analysis["notes"].append("Contains water quality test data")
        if "compliance" in sheet_name.lower():
            sheet_analysis["notes"].append("Contains compliance reporting data")

        analysis["sheets"].append(sheet_analysis)

    # Save analysis
    with open(OUTPUT_PATH, "w") as f:
        json.dump(analysis, f, indent=2)

    print()
    print("=" * 70)
    print("ANALYSIS COMPLETE")
    print("=" * 70)
    print(f"\nField mapping saved to: {OUTPUT_PATH}")
    print(f"\nNext steps:")
    print("  1. Review the header names and map to database fields")
    print("  2. Identify required vs optional fields")
    print("  3. Update database schema if needed")
    print("  4. Implement export service")
    print()

if __name__ == "__main__":
    analyze_excel_template()
