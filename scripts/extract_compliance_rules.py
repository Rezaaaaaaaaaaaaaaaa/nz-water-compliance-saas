"""
Extract Compliance Rules from DWQAR Excel Template
Extracts all 381 Rule IDs from the RuleIDs sheet and generates JSON for database seeding
"""

import openpyxl
import json
from pathlib import Path

TEMPLATE_PATH = Path("C:/compliance-saas/docs/regulations/taumata-arowai/DWQAR_Reporting_Template.xlsx")
OUTPUT_PATH = Path("C:/compliance-saas/backend/prisma/seeds/compliance_rules.json")

def extract_rules():
    """Extract all compliance rules from the RuleIDs sheet"""

    print("=" * 70)
    print("COMPLIANCE RULES EXTRACTION")
    print("=" * 70)
    print()

    if not TEMPLATE_PATH.exists():
        print(f"[ERROR] Template not found: {TEMPLATE_PATH}")
        return

    # Ensure output directory exists
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    # Load workbook
    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    if "RuleIDs" not in wb.sheetnames:
        print("[ERROR] RuleIDs sheet not found in template")
        return

    sheet = wb["RuleIDs"]
    print(f"Sheet: RuleIDs")
    print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns\n")

    # Extract all rule IDs (column A)
    rules = []
    rule_ids_seen = set()

    for row_idx in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_idx, column=1).value

        if cell_value and str(cell_value).strip():
            rule_id = str(cell_value).strip()

            # Skip duplicates
            if rule_id in rule_ids_seen:
                continue

            rule_ids_seen.add(rule_id)

            # Parse rule ID to extract category and parameter
            category = None
            parameter = None

            # Rule ID format: T1.8-ecol, T2.1-pH, etc.
            if '-' in rule_id:
                parts = rule_id.split('-')
                if len(parts) == 2:
                    parameter = parts[1].lower()

            # Categorize based on rule prefix
            if rule_id.startswith('T1'):
                category = "BACTERIOLOGICAL"
            elif rule_id.startswith('T2'):
                category = "CHEMICAL"
            elif rule_id.startswith('T3'):
                category = "PROTOZOA"
            elif rule_id.startswith('T4'):
                category = "RADIOLOGICAL"
            elif rule_id.startswith('M'):
                category = "MONITORING"
            elif rule_id.startswith('V'):
                category = "VERIFICATION"
            elif rule_id.startswith('O'):
                category = "OPERATIONAL"
            else:
                category = "WATER_QUALITY"

            # Build rule object
            rule = {
                "ruleId": rule_id,
                "category": category,
                "parameter": parameter,
                "description": f"Compliance rule {rule_id}",
                "isActive": True,
                "applicability": "All supply sizes",
                "effectiveDate": "2024-01-01T00:00:00.000Z"
            }

            rules.append(rule)

            if len(rules) % 50 == 0:
                print(f"Extracted {len(rules)} rules...")

    print()
    print(f"[OK] Total rules extracted: {len(rules)}")
    print()

    # Show sample rules
    print("Sample rules:")
    for rule in rules[:10]:
        print(f"  {rule['ruleId']} - {rule['category']} ({rule['parameter'] or 'N/A'})")
    print()

    # Save to JSON
    output_data = {
        "metadata": {
            "source": "DWQAR_Reporting_Template.xlsx",
            "sheet": "RuleIDs",
            "extractionDate": "2025-10-05",
            "totalRules": len(rules)
        },
        "rules": rules
    }

    with open(OUTPUT_PATH, "w") as f:
        json.dump(output_data, f, indent=2)

    print(f"[OK] Rules saved to: {OUTPUT_PATH}")
    print()

    # Category breakdown
    category_counts = {}
    for rule in rules:
        cat = rule['category']
        category_counts[cat] = category_counts.get(cat, 0) + 1

    print("Rules by category:")
    for cat, count in sorted(category_counts.items()):
        print(f"  {cat}: {count}")
    print()

    print("=" * 70)
    print("EXTRACTION COMPLETE")
    print("=" * 70)
    print()
    print("Next step: Create Prisma seed script to import these rules")
    print()

if __name__ == "__main__":
    extract_rules()
